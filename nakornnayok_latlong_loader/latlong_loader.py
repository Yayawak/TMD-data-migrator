from openpyxl import Workbook, load_workbook
from typing import List, Tuple
import numpy as np

# def getLatLongOfNakornnayok() -> List[Tuple[str, List[List]]]:
def getLatLongOfNakornnayok() -> List[Tuple[str, List[List]]]:
# def getLatLongOfNakornnayok() -> np.ndarray[str, np.ndarray[np.ndarray]]:
    # wb = Workbook()
    nakornnayok_wb = load_workbook("resources/นครนายก.xlsx")
    sheet_data = []
    for i, ws in enumerate(nakornnayok_wb.worksheets):
        sheet_name = nakornnayok_wb.sheetnames[i]
        print(f"sheet {i} : {sheet_name}")


#NOTE - structure should like this ...
# [
#   ['sheet0', 2d-lat-long],
#   ['sheet1', 2d-lat-long],
#   ['sheet2', 2d-lat-long],
#   ['sheet3', 2d-lat-long],
# ]
        latlongs = []

        for row in ws.iter_rows(min_row=2, max_col=2):
            # print(row)
            lat = row[0].value
            long = row[1].value

            # A = np.array([lat, long])
            # latlongs.append(A)
            latlongs.append([lat, long])
            # print(latlongs)
            print([lat, long])
            # print()
        # B = np.array([sheet_name, np.array(latlongs)])
        # sheet_data.append(B)
        sheet_data.append([sheet_name, latlongs])
        print("-----------")
    
    return sheet_data
    
    